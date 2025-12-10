"""
Document Upload and OCR Processing API Routes.
Handles customer document uploads, OCR extraction, and agent notifications.
"""

from fastapi import APIRouter, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
from datetime import datetime
import os
import uuid
import json
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter()

# Storage paths
UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads")
DOCUMENTS_DIR = os.path.join(UPLOAD_DIR, "documents")
PROCESSED_DIR = os.path.join(UPLOAD_DIR, "processed")
CRM_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "crm")

# Ensure directories exist
for dir_path in [UPLOAD_DIR, DOCUMENTS_DIR, PROCESSED_DIR, CRM_DIR]:
    os.makedirs(dir_path, exist_ok=True)

# In-memory storage for notifications (in production, use Redis/WebSocket)
agent_notifications: Dict[str, List[Dict]] = {}
document_status: Dict[str, Dict] = {}


class DocumentUploadResponse(BaseModel):
    """Response model for document upload."""
    success: bool
    document_id: str
    filename: str
    document_type: Optional[str]
    status: str
    message: str
    extracted_data: Optional[Dict[str, Any]] = None


class CustomerDocumentRequest(BaseModel):
    """Request model for customer document submission."""
    customer_name: str
    customer_email: str
    customer_phone: str
    policy_number: Optional[str] = None
    agent_code: Optional[str] = None
    notes: Optional[str] = None


class AgentNotification(BaseModel):
    """Model for agent notifications."""
    notification_id: str
    customer_name: str
    document_type: str
    document_id: str
    timestamp: str
    status: str
    extracted_data: Optional[Dict[str, Any]] = None


def process_document_async(document_id: str, file_path: str, customer_info: dict, agent_code: str):
    """
    Background task to process uploaded document with OCR.
    Updates CRM and sends notification to agent.
    """
    try:
        # Import OCR engine
        import sys
        ml_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "ml")
        sys.path.insert(0, ml_path)
        
        from vision.ocr_engine import InsuranceDocumentOCR
        from vision.excel_crm import ExcelCRMManager
        
        # Initialize OCR
        ocr = InsuranceDocumentOCR()
        
        # Process document
        logger.info(f"Processing document {document_id}: {file_path}")
        result = ocr.process_document(file_path)
        
        # Update document status
        document_status[document_id] = {
            "status": "processed" if result.get("success") else "failed",
            "document_type": result.get("document_type", "unknown"),
            "extracted_data": result.get("extracted_data", {}),
            "confidence": result.get("confidence", 0),
            "processed_at": datetime.now().isoformat()
        }
        
        # Update Excel CRM
        crm = ExcelCRMManager(CRM_DIR)
        crm.add_document_record(
            agent_code=agent_code or "UNASSIGNED",
            customer_info=customer_info,
            document_info={
                "document_id": document_id,
                "document_type": result.get("document_type", "unknown"),
                "file_path": file_path,
                "extracted_data": result.get("extracted_data", {}),
                "confidence": result.get("confidence", 0),
                "uploaded_at": customer_info.get("uploaded_at", datetime.now().isoformat())
            }
        )
        
        # Create agent notification
        notification = {
            "notification_id": str(uuid.uuid4()),
            "customer_name": customer_info.get("customer_name", "Unknown"),
            "customer_email": customer_info.get("customer_email", ""),
            "customer_phone": customer_info.get("customer_phone", ""),
            "document_type": result.get("document_type", "unknown"),
            "document_id": document_id,
            "timestamp": datetime.now().isoformat(),
            "status": "new",
            "extracted_data": result.get("extracted_data", {}),
            "confidence": result.get("confidence", 0)
        }
        
        # Add to agent notifications queue
        if agent_code not in agent_notifications:
            agent_notifications[agent_code] = []
        agent_notifications[agent_code].append(notification)
        
        # Also notify unassigned queue for routing
        if agent_code != "UNASSIGNED":
            if "UNASSIGNED" not in agent_notifications:
                agent_notifications["UNASSIGNED"] = []
        
        logger.info(f"Document {document_id} processed successfully. Type: {result.get('document_type')}")
        
        # Move to processed folder
        processed_path = os.path.join(PROCESSED_DIR, os.path.basename(file_path))
        os.rename(file_path, processed_path)
        
    except Exception as e:
        logger.error(f"Error processing document {document_id}: {str(e)}")
        document_status[document_id] = {
            "status": "error",
            "error": str(e),
            "processed_at": datetime.now().isoformat()
        }


@router.post("/upload", response_model=DocumentUploadResponse)
async def upload_document(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    customer_name: str = Form(...),
    customer_email: str = Form(...),
    customer_phone: str = Form(...),
    policy_number: Optional[str] = Form(None),
    agent_code: Optional[str] = Form(None),
    notes: Optional[str] = Form(None)
):
    """
    Upload a customer document (PDF or image) for OCR processing.
    
    The document will be:
    1. Saved to the upload directory
    2. Processed via OCR in the background
    3. Classified and data extracted
    4. Added to the agent's Excel CRM
    5. Agent notified of new document
    """
    # Validate file type
    allowed_extensions = {'.pdf', '.png', '.jpg', '.jpeg', '.tiff', '.bmp'}
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    if file_ext not in allowed_extensions:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
        )
    
    # Generate unique document ID
    document_id = f"DOC-{datetime.now().strftime('%Y%m%d')}-{str(uuid.uuid4())[:8].upper()}"
    
    # Create filename with document ID
    safe_filename = f"{document_id}{file_ext}"
    file_path = os.path.join(DOCUMENTS_DIR, safe_filename)
    
    try:
        # Save uploaded file
        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)
        
        # Initialize document status
        document_status[document_id] = {
            "status": "processing",
            "uploaded_at": datetime.now().isoformat(),
            "original_filename": file.filename
        }
        
        # Customer info for CRM
        customer_info = {
            "customer_name": customer_name,
            "customer_email": customer_email,
            "customer_phone": customer_phone,
            "policy_number": policy_number,
            "notes": notes,
            "uploaded_at": datetime.now().isoformat()
        }
        
        # Schedule background processing
        background_tasks.add_task(
            process_document_async,
            document_id,
            file_path,
            customer_info,
            agent_code or "UNASSIGNED"
        )
        
        return DocumentUploadResponse(
            success=True,
            document_id=document_id,
            filename=file.filename,
            document_type=None,  # Will be determined by OCR
            status="processing",
            message="Document uploaded successfully. Processing in background."
        )
        
    except Exception as e:
        logger.error(f"Error uploading document: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@router.get("/status/{document_id}")
async def get_document_status(document_id: str):
    """Get the processing status of a document."""
    if document_id not in document_status:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return {
        "document_id": document_id,
        **document_status[document_id]
    }


@router.get("/notifications/{agent_code}")
async def get_agent_notifications(agent_code: str, mark_read: bool = False):
    """
    Get pending notifications for an agent.
    
    Args:
        agent_code: The agent's unique code
        mark_read: If True, mark all notifications as read
    """
    notifications = agent_notifications.get(agent_code, [])
    
    # Filter to only unread/new notifications
    pending = [n for n in notifications if n.get("status") == "new"]
    
    if mark_read:
        for n in notifications:
            n["status"] = "read"
    
    return {
        "agent_code": agent_code,
        "total_notifications": len(pending),
        "notifications": pending
    }


@router.post("/notifications/{agent_code}/clear")
async def clear_agent_notifications(agent_code: str):
    """Clear all notifications for an agent."""
    if agent_code in agent_notifications:
        agent_notifications[agent_code] = []
    
    return {"success": True, "message": "Notifications cleared"}


@router.get("/customer/{customer_email}")
async def get_customer_documents(customer_email: str):
    """Get all documents for a customer by email."""
    # Search through all agent CRM files
    customer_docs = []
    
    for filename in os.listdir(CRM_DIR):
        if filename.endswith('.xlsx'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(os.path.join(CRM_DIR, filename))
                
                for sheet_name in wb.sheetnames:
                    if sheet_name == "Documents":
                        sheet = wb[sheet_name]
                        headers = [cell.value for cell in sheet[1]]
                        
                        if "Customer Email" in headers:
                            email_idx = headers.index("Customer Email")
                            
                            for row in sheet.iter_rows(min_row=2, values_only=True):
                                if row[email_idx] == customer_email:
                                    doc = dict(zip(headers, row))
                                    customer_docs.append(doc)
                
            except Exception as e:
                logger.warning(f"Error reading CRM file {filename}: {str(e)}")
    
    return {
        "customer_email": customer_email,
        "total_documents": len(customer_docs),
        "documents": customer_docs
    }


@router.get("/types")
async def get_supported_document_types():
    """Get list of supported document types for upload."""
    return {
        "document_types": [
            {
                "type": "logbook",
                "name": "Vehicle Logbook",
                "description": "Kenya motor vehicle registration document",
                "extracted_fields": ["registration_number", "owner_name", "chassis_number", "engine_number", "make_model", "year_of_manufacture"]
            },
            {
                "type": "national_id",
                "name": "National ID",
                "description": "Kenya National ID card",
                "extracted_fields": ["id_number", "full_name", "date_of_birth", "gender", "district", "serial_number"]
            },
            {
                "type": "driving_license",
                "name": "Driving License",
                "description": "Kenya Driving License",
                "extracted_fields": ["license_number", "holder_name", "license_class", "issue_date", "expiry_date"]
            },
            {
                "type": "insurance_proposal",
                "name": "Insurance Proposal Form",
                "description": "Motor vehicle insurance proposal form",
                "extracted_fields": ["proposer_name", "vehicle_details", "coverage_requested"]
            },
            {
                "type": "vehicle_photo",
                "name": "Vehicle Photo",
                "description": "Photo of the insured vehicle",
                "extracted_fields": ["visual_inspection"]
            }
        ],
        "accepted_formats": ["PDF", "PNG", "JPG", "JPEG", "TIFF", "BMP"]
    }


@router.get("/stats")
async def get_document_stats():
    """Get overall document processing statistics."""
    total = len(document_status)
    processed = sum(1 for d in document_status.values() if d.get("status") == "processed")
    processing = sum(1 for d in document_status.values() if d.get("status") == "processing")
    failed = sum(1 for d in document_status.values() if d.get("status") in ["error", "failed"])
    
    # Count by document type
    type_counts = {}
    for doc in document_status.values():
        doc_type = doc.get("document_type", "unknown")
        type_counts[doc_type] = type_counts.get(doc_type, 0) + 1
    
    return {
        "total_documents": total,
        "processed": processed,
        "processing": processing,
        "failed": failed,
        "success_rate": round(processed / total * 100, 1) if total > 0 else 0,
        "by_type": type_counts
    }
