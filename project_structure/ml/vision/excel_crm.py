"""
Excel CRM Manager for Insurance Document Management.
Manages agent-specific Excel workbooks for tracking customer documents.

Each agent gets their own Excel workbook with sheets for:
- Customers: Customer information
- Documents: Uploaded documents with extracted data
- Policies: Policy information
- Activity Log: All actions and updates
"""

import os
import json
from datetime import datetime
from typing import Dict, Any, List, Optional
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExcelCRMManager:
    """
    Manages Excel-based CRM for insurance agents.
    Each agent has their own workbook with customer and document data.
    """
    
    def __init__(self, crm_directory: str):
        """
        Initialize the CRM manager.
        
        Args:
            crm_directory: Directory where agent CRM files are stored
        """
        self.crm_directory = crm_directory
        os.makedirs(crm_directory, exist_ok=True)
        
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not installed. Excel CRM features disabled.")
    
    def _get_agent_file(self, agent_code: str) -> str:
        """Get the Excel file path for an agent."""
        return os.path.join(self.crm_directory, f"CRM_{agent_code}.xlsx")
    
    def _create_workbook(self, agent_code: str) -> 'openpyxl.Workbook':
        """Create a new CRM workbook for an agent."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel CRM features")
        
        wb = openpyxl.Workbook()
        
        # Style definitions
        header_fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Customers
        ws_customers = wb.active
        ws_customers.title = "Customers"
        customer_headers = [
            "Customer ID", "Customer Name", "Email", "Phone", 
            "Policy Number", "Registration Date", "Status", "Notes"
        ]
        for col, header in enumerate(customer_headers, 1):
            cell = ws_customers.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            ws_customers.column_dimensions[get_column_letter(col)].width = 18
        
        # Sheet 2: Documents
        ws_docs = wb.create_sheet("Documents")
        doc_headers = [
            "Document ID", "Customer Name", "Customer Email", "Customer Phone",
            "Document Type", "File Path", "Upload Date", "Process Status",
            "Confidence Score", "Extracted Data", "Agent Notes"
        ]
        for col, header in enumerate(doc_headers, 1):
            cell = ws_docs.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            ws_docs.column_dimensions[get_column_letter(col)].width = 20
        
        # Wider column for extracted data
        ws_docs.column_dimensions['J'].width = 50
        
        # Sheet 3: Policies
        ws_policies = wb.create_sheet("Policies")
        policy_headers = [
            "Policy Number", "Customer Name", "Vehicle Registration", 
            "Coverage Type", "Start Date", "End Date", "Premium",
            "Status", "Insurer", "Commission"
        ]
        for col, header in enumerate(policy_headers, 1):
            cell = ws_policies.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            ws_policies.column_dimensions[get_column_letter(col)].width = 18
        
        # Sheet 4: Activity Log
        ws_log = wb.create_sheet("Activity Log")
        log_headers = [
            "Timestamp", "Activity Type", "Customer", "Document ID",
            "Description", "Status"
        ]
        for col, header in enumerate(log_headers, 1):
            cell = ws_log.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            ws_log.column_dimensions[get_column_letter(col)].width = 20
        ws_log.column_dimensions['E'].width = 50  # Description column wider
        
        # Sheet 5: Summary Dashboard
        ws_summary = wb.create_sheet("Summary")
        summary_headers = ["Metric", "Value", "Last Updated"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Add initial metrics
        metrics = [
            ("Total Customers", 0),
            ("Total Documents", 0),
            ("Active Policies", 0),
            ("Documents Pending Review", 0),
            ("This Month New Customers", 0),
            ("This Month Documents", 0)
        ]
        for row, (metric, value) in enumerate(metrics, 2):
            ws_summary.cell(row=row, column=1, value=metric)
            ws_summary.cell(row=row, column=2, value=value)
            ws_summary.cell(row=row, column=3, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 20
        
        return wb
    
    def _get_or_create_workbook(self, agent_code: str) -> 'openpyxl.Workbook':
        """Get existing workbook or create new one."""
        file_path = self._get_agent_file(agent_code)
        
        if os.path.exists(file_path):
            return openpyxl.load_workbook(file_path)
        else:
            return self._create_workbook(agent_code)
    
    def add_document_record(
        self,
        agent_code: str,
        customer_info: Dict[str, Any],
        document_info: Dict[str, Any]
    ) -> bool:
        """
        Add a document record to the agent's CRM.
        
        Args:
            agent_code: Agent's unique code
            customer_info: Customer details (name, email, phone, etc.)
            document_info: Document details (id, type, extracted_data, etc.)
            
        Returns:
            bool: Success status
        """
        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl not installed. Cannot update CRM.")
            return False
        
        try:
            wb = self._get_or_create_workbook(agent_code)
            
            # Add to Documents sheet
            ws_docs = wb["Documents"]
            next_row = ws_docs.max_row + 1
            
            # Prepare extracted data as JSON string
            extracted_data = document_info.get("extracted_data", {})
            extracted_str = json.dumps(extracted_data, indent=2) if extracted_data else ""
            
            # Add document row
            doc_row = [
                document_info.get("document_id", ""),
                customer_info.get("customer_name", ""),
                customer_info.get("customer_email", ""),
                customer_info.get("customer_phone", ""),
                document_info.get("document_type", "unknown"),
                document_info.get("file_path", ""),
                document_info.get("uploaded_at", datetime.now().isoformat()),
                "Processed" if document_info.get("confidence", 0) > 0 else "Pending",
                f"{document_info.get('confidence', 0):.1%}",
                extracted_str,
                customer_info.get("notes", "")
            ]
            
            for col, value in enumerate(doc_row, 1):
                ws_docs.cell(row=next_row, column=col, value=value)
            
            # Check if customer exists, if not add to Customers sheet
            ws_customers = wb["Customers"]
            customer_exists = False
            customer_email = customer_info.get("customer_email", "")
            
            for row in ws_customers.iter_rows(min_row=2, max_col=3, values_only=True):
                if row[2] == customer_email:  # Email column
                    customer_exists = True
                    break
            
            if not customer_exists and customer_email:
                next_cust_row = ws_customers.max_row + 1
                customer_id = f"CUST-{datetime.now().strftime('%Y%m%d')}-{next_cust_row:04d}"
                customer_row = [
                    customer_id,
                    customer_info.get("customer_name", ""),
                    customer_email,
                    customer_info.get("customer_phone", ""),
                    customer_info.get("policy_number", ""),
                    datetime.now().strftime("%Y-%m-%d"),
                    "Active",
                    customer_info.get("notes", "")
                ]
                for col, value in enumerate(customer_row, 1):
                    ws_customers.cell(row=next_cust_row, column=col, value=value)
            
            # Add to Activity Log
            ws_log = wb["Activity Log"]
            next_log_row = ws_log.max_row + 1
            log_row = [
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Document Upload",
                customer_info.get("customer_name", ""),
                document_info.get("document_id", ""),
                f"New {document_info.get('document_type', 'document')} uploaded. "
                f"Confidence: {document_info.get('confidence', 0):.1%}",
                "Completed"
            ]
            for col, value in enumerate(log_row, 1):
                ws_log.cell(row=next_log_row, column=col, value=value)
            
            # Update Summary metrics
            self._update_summary(wb)
            
            # Save workbook
            file_path = self._get_agent_file(agent_code)
            wb.save(file_path)
            logger.info(f"Document record added to CRM for agent {agent_code}")
            
            return True
            
        except Exception as e:
            logger.error(f"Error adding document to CRM: {str(e)}")
            return False
    
    def _update_summary(self, wb: 'openpyxl.Workbook'):
        """Update the summary dashboard metrics."""
        ws_summary = wb["Summary"]
        
        # Count customers
        ws_customers = wb["Customers"]
        total_customers = max(0, ws_customers.max_row - 1)
        
        # Count documents
        ws_docs = wb["Documents"]
        total_docs = max(0, ws_docs.max_row - 1)
        
        # Count pending documents
        pending = 0
        for row in ws_docs.iter_rows(min_row=2, max_col=8, values_only=True):
            if row[7] == "Pending":
                pending += 1
        
        # Count this month's entries
        current_month = datetime.now().strftime("%Y-%m")
        this_month_customers = 0
        this_month_docs = 0
        
        for row in ws_customers.iter_rows(min_row=2, max_col=6, values_only=True):
            if row[5] and str(row[5]).startswith(current_month):
                this_month_customers += 1
        
        for row in ws_docs.iter_rows(min_row=2, max_col=7, values_only=True):
            if row[6] and str(row[6]).startswith(current_month):
                this_month_docs += 1
        
        # Update summary values
        metrics = [
            total_customers,
            total_docs,
            0,  # Active policies - would need policy logic
            pending,
            this_month_customers,
            this_month_docs
        ]
        
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        for row, value in enumerate(metrics, 2):
            ws_summary.cell(row=row, column=2, value=value)
            ws_summary.cell(row=row, column=3, value=now)
    
    def get_agent_summary(self, agent_code: str) -> Dict[str, Any]:
        """Get summary statistics for an agent."""
        if not OPENPYXL_AVAILABLE:
            return {"error": "openpyxl not installed"}
        
        file_path = self._get_agent_file(agent_code)
        if not os.path.exists(file_path):
            return {"error": "No CRM file found for agent"}
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws_summary = wb["Summary"]
            
            summary = {}
            for row in ws_summary.iter_rows(min_row=2, max_row=7, values_only=True):
                if row[0]:
                    summary[row[0]] = row[1]
            
            return {
                "agent_code": agent_code,
                "summary": summary,
                "file_path": file_path
            }
            
        except Exception as e:
            return {"error": str(e)}
    
    def get_recent_documents(self, agent_code: str, limit: int = 10) -> List[Dict[str, Any]]:
        """Get recent documents for an agent."""
        if not OPENPYXL_AVAILABLE:
            return []
        
        file_path = self._get_agent_file(agent_code)
        if not os.path.exists(file_path):
            return []
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws_docs = wb["Documents"]
            
            headers = [cell.value for cell in ws_docs[1]]
            documents = []
            
            # Get all rows and sort by date (newest first)
            for row in ws_docs.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Has document ID
                    doc = dict(zip(headers, row))
                    documents.append(doc)
            
            # Sort by upload date descending
            documents.sort(
                key=lambda x: x.get("Upload Date", "") or "", 
                reverse=True
            )
            
            return documents[:limit]
            
        except Exception as e:
            logger.error(f"Error getting recent documents: {str(e)}")
            return []
    
    def search_documents(
        self, 
        agent_code: str, 
        search_term: str,
        field: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Search documents in agent's CRM.
        
        Args:
            agent_code: Agent's unique code
            search_term: Term to search for
            field: Specific field to search (optional)
            
        Returns:
            List of matching documents
        """
        if not OPENPYXL_AVAILABLE:
            return []
        
        file_path = self._get_agent_file(agent_code)
        if not os.path.exists(file_path):
            return []
        
        try:
            wb = openpyxl.load_workbook(file_path)
            ws_docs = wb["Documents"]
            
            headers = [cell.value for cell in ws_docs[1]]
            matches = []
            search_lower = search_term.lower()
            
            for row in ws_docs.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Has document ID
                    doc = dict(zip(headers, row))
                    
                    if field and field in doc:
                        # Search specific field
                        if search_lower in str(doc[field]).lower():
                            matches.append(doc)
                    else:
                        # Search all fields
                        for value in row:
                            if value and search_lower in str(value).lower():
                                matches.append(doc)
                                break
            
            return matches
            
        except Exception as e:
            logger.error(f"Error searching documents: {str(e)}")
            return []


# Standalone test
if __name__ == "__main__":
    # Test the CRM manager
    crm = ExcelCRMManager("./test_crm")
    
    # Add a test document
    success = crm.add_document_record(
        agent_code="AGT001",
        customer_info={
            "customer_name": "John Kamau",
            "customer_email": "john.kamau@email.com",
            "customer_phone": "+254712345678",
            "policy_number": "POL-2024-001",
            "notes": "New customer referral"
        },
        document_info={
            "document_id": "DOC-20240101-ABCD1234",
            "document_type": "logbook",
            "file_path": "/uploads/DOC-20240101-ABCD1234.pdf",
            "extracted_data": {
                "registration_number": "KDA 123A",
                "owner_name": "John Kamau",
                "chassis_number": "ABC123456789"
            },
            "confidence": 0.92,
            "uploaded_at": datetime.now().isoformat()
        }
    )
    
    print(f"Document added: {success}")
    print(f"Summary: {crm.get_agent_summary('AGT001')}")
    print(f"Recent docs: {crm.get_recent_documents('AGT001')}")
